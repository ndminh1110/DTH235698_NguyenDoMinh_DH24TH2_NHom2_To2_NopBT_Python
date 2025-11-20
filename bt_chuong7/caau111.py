from openpyxl import Workbook, load_workbook
import os

FILENAME = "nhanvien.xlsx"

def tao_file_moi():
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"
    ws.append(["STT", "Mã", "Tên", "Tuổi"])
    wb.save(FILENAME)
    print("Đã tạo file Excel mới.")

def them_nv():
    ma = input("Nhập mã NV: ")
    ten = input("Nhập tên NV: ")
    tuoi = int(input("Nhập tuổi NV: "))

    if not os.path.exists(FILENAME):
        tao_file_moi()

    wb = load_workbook(FILENAME)
    ws = wb.active
    stt = ws.max_row  # đếm hàng hiện tại
    ws.append([stt, ma, ten, tuoi])
    wb.save(FILENAME)
    print("Đã thêm nhân viên vào Excel.")

def doc_ds():
    if not os.path.exists(FILENAME):
        print("Chưa có file Excel.")
        return
    wb = load_workbook(FILENAME)
    ws = wb.active
    print("\nDanh sách nhân viên:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)
    wb.close()

def sapxep_theo_tuoi():
    if not os.path.exists(FILENAME):
        print("Chưa có file Excel.")
        return
    wb = load_workbook(FILENAME)
    ws = wb.active
    data = list(ws.iter_rows(min_row=2, values_only=True))
    data.sort(key=lambda x: x[3])  # sắp xếp theo cột Tuổi

    ws.delete_rows(2, ws.max_row)
    for i, row in enumerate(data, start=1):
        ws.append([i] + list(row[1:]))
    wb.save(FILENAME)
    print("Đã sắp xếp theo tuổi tăng dần.")

def main():
    while True:
        print("\n=== QUẢN LÝ NHÂN VIÊN (Excel) ===")
        print("1. Tạo file Excel mới")
        print("2. Thêm nhân viên")
        print("3. Xem danh sách")
        print("4. Sắp xếp theo tuổi")
        print("5. Thoát")
        chon = input("Chọn chức năng (1-5): ")

        if chon == "1": tao_file_moi()
        elif chon == "2": them_nv()
        elif chon == "3": doc_ds()
        elif chon == "4": sapxep_theo_tuoi()
        elif chon == "5":
            print("Kết thúc chương trình.")
            break
        else:
            print("Lựa chọn không hợp lệ!")

if __name__ == "__main__":
    main()